# ライブラリ
import openpyxl
import datetime
import pandas as pd
import shutil

# 時間を取得
today = datetime.datetime.today()
time_now = datetime.datetime.strftime(today, '%H:%M')
day_today = today.day
day_month = today.month

#　--- main ---
# ecxelの作成
if day_today == 1:
    shutil.copy('./タイムカード(月)_上野.xlsx','./タイムカード('+str(day_month)+'月)_上野.xlsx')
else:
    pass    

# excelを開く
wb = openpyxl.load_workbook('./タイムカード('+str(day_month)+'月)_上野.xlsx',data_only=True)
sheet = wb['Table 1']

# 表題
if  day_today == 1:
    today_format = "{0:%Y/%m/%d}".format(today)
    sheet.cell(row=1,column=1).value =today_format
    wb.save('./タイムカード('+str(day_month)+'月)_上野.xlsx')
else:
    pass

# 該当セルの行番号を取得
for cells in sheet['A7':'A37']:
    for test_cell in cells:
        test_day = test_cell.value
        test_day = pd.to_timedelta(test_day,unit='D')+pd.to_datetime("1899/12/30")
        if  test_day.day == day_today:
            a = test_cell.row
            break
    else:
        continue    
    break

# 時刻記入
b = sheet.cell(row=a,column=6).coordinate
if sheet[b].value == None:
    sheet[b].value = time_now
    wb.save('./タイムカード('+str(day_month)+'月)_上野.xlsx')
elif sheet[b].value != None:
    c = sheet.cell(row=a,column=7).coordinate
    sheet[c].value = time_now
    wb.save('./タイムカード('+str(day_month)+'月)_上野.xlsx')

# ecxelを閉じる
wb.close()
